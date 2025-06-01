from fastapi import APIRouter, Depends, HTTPException, Response, Query  # ✅ MAKE SURE Response IS HERE
from fastapi.responses import FileResponse  # ✅ Alternative import
from typing import List, Optional
from models.user import User
from models.table import Table
from models.pdf import PDF
from auth.dependencies import get_current_active_user
from utils.pydantic_objectid import PyObjectId
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

router = APIRouter(prefix="/tables", tags=["Tables"])

@router.get("/document/{document_id}")
async def get_document_tables(
    document_id: str,
    page: Optional[int] = Query(1, ge=1),
    limit: Optional[int] = Query(10, ge=1, le=50),
    search: Optional[str] = Query(None),
    current_user: User = Depends(get_current_active_user)
):
    """Get all tables from a specific document with pagination and search"""
    try:
        # Verify document exists and user has access
        document = await PDF.get(PyObjectId(document_id))
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        if document.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Get ALL tables first, then filter in Python
        all_tables = await Table.find(Table.pdf_id == PyObjectId(document_id)).to_list()
        
        # Apply search filter in Python if needed
        if search:
            search_lower = search.lower()
            filtered_tables = [
                table for table in all_tables 
                if search_lower in (getattr(table, 'table_title', '') or '').lower() or 
                   search_lower in (getattr(table, 'markdown_content', '') or '').lower()
            ]
        else:
            filtered_tables = all_tables
        
        # Sort by table number
        filtered_tables.sort(key=lambda x: getattr(x, 'table_number', 0))
        
        # Apply pagination in Python
        total_tables = len(filtered_tables)
        skip = (page - 1) * limit
        paginated_tables = filtered_tables[skip:skip + limit]
        
        # Format response
        table_list = []
        for table in paginated_tables:
            table_data = {
                "id": str(table.id),
                "title": getattr(table, 'table_title', f'Table {getattr(table, "table_number", 1)}'),
                "table_number": getattr(table, 'table_number', 1),
                "start_page": getattr(table, 'start_page', 1),
                "end_page": getattr(table, 'end_page', getattr(table, 'start_page', 1)),
                "column_count": getattr(table, 'column_count', 0),
                "row_count": getattr(table, 'row_count', 0),
                "markdown_content": getattr(table, 'markdown_content', ''),
            }
            table_list.append(table_data)
        
        return {
            "success": True,
            "tables": table_list,
            "pagination": {
                "page": page,
                "limit": limit,
                "total": total_tables,
                "pages": (total_tables + limit - 1) // limit if total_tables > 0 else 0
            },
            "document": {
                "id": str(document.id),
                "filename": document.filename,
                "total_tables": total_tables
            }
        }
        
    except ValueError as e:
        print(f"ValueError in get_document_tables: {e}")
        raise HTTPException(status_code=400, detail="Invalid document ID")
    except Exception as e:
        print(f"Error in get_document_tables: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@router.get("/document/{document_id}/summary")
async def get_document_tables_summary(
    document_id: str,
    current_user: User = Depends(get_current_active_user)
):
    """Get summary of tables in a document"""
    try:
        # Verify document exists and user has access
        document = await PDF.get(PyObjectId(document_id))
        if not document:
            raise HTTPException(status_code=404, detail="Document not found")
        
        if document.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Get all tables
        tables = await Table.find(Table.pdf_id == PyObjectId(document_id)).to_list()
        
        # Calculate summary stats
        total_tables = len(tables)
        total_rows = sum(getattr(table, 'row_count', 0) for table in tables)
        total_columns = sum(getattr(table, 'column_count', 0) for table in tables)
        avg_columns = total_columns / total_tables if total_tables > 0 else 0
        
        # Get pages that contain tables
        pages_with_tables = list(set(getattr(table, 'start_page', 0) for table in tables if hasattr(table, 'start_page')))
        pages_with_tables.sort()
        
        return {
            "success": True,
            "summary": {
                "total_tables": total_tables,
                "total_rows": total_rows,
                "total_columns": total_columns,
                "average_columns": round(avg_columns, 1),
                "pages_with_tables": pages_with_tables,
                "page_count": len(pages_with_tables)
            },
            "document": {
                "id": str(document.id),
                "filename": document.filename,
                "page_count": getattr(document, 'page_count', 0)
            }
        }
        
    except ValueError as e:
        print(f"ValueError in get_document_tables_summary: {e}")
        raise HTTPException(status_code=400, detail="Invalid document ID")
    except Exception as e:
        print(f"Error in get_document_tables_summary: {e}")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@router.post("/{table_id}/export")
async def export_single_table(
    table_id: str,
    current_user: User = Depends(get_current_active_user)
):
    """Export a single table as Excel"""
    try:
        print(f"Starting export for table_id: {table_id}")
        
        # Get the table
        table = await Table.get(PyObjectId(table_id))
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")
        
        print(f"Found table: {getattr(table, 'table_title', 'Unknown')}")
        
        # Check if user owns the document containing this table
        document = await PDF.get(table.pdf_id)
        if not document or document.user_id != current_user.id:
            raise HTTPException(status_code=403, detail="Access denied")
        
        print(f"User has access to document: {document.filename}")
        
        # Parse markdown table to DataFrame
        markdown_content = getattr(table, 'markdown_content', '')
        print(f"Markdown content length: {len(markdown_content)}")
        
        df = parse_markdown_table_to_dataframe(markdown_content)
        
        if df is None or df.empty:
            print("DataFrame is empty or None")
            raise HTTPException(status_code=400, detail="Unable to parse table data")
        
        print(f"DataFrame created with shape: {df.shape}")
        
        # Create Excel workbook with single sheet
        wb = Workbook()
        ws = wb.active
        
        table_title = getattr(table, 'table_title', f'Table {getattr(table, "table_number", 1)}')
        ws.title = table_title[:31] if table_title else "Table"
        
        # Add table info header
        ws.append([f"Table: {table_title}"])
        ws.append([f"Page: {getattr(table, 'start_page', 1)}" + 
                  (f"-{getattr(table, 'end_page', getattr(table, 'start_page', 1))}" 
                   if getattr(table, 'start_page', 1) != getattr(table, 'end_page', getattr(table, 'start_page', 1)) else "")])
        ws.append([f"Rows: {getattr(table, 'row_count', 0)} | Columns: {getattr(table, 'column_count', 0)}"])
        ws.append([f"From Document: {document.filename}"])
        ws.append([])  # Empty row
        
        # Style header rows
        for row in range(1, 5):
            cell = ws.cell(row=row, column=1)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add column headers
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_idx, value=str(header))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_idx, row_data in enumerate(df.values, 7):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print("Excel workbook created successfully")
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        table_name = table_title.replace(' ', '_') if table_title else f"table_{getattr(table, 'table_number', 1)}"
        # Clean filename
        table_name = re.sub(r'[^\w\-_\.]', '_', table_name)
        filename = f"{table_name}.xlsx"
        
        print(f"Returning Excel file: {filename}")
        
        # ✅ FIXED: Return Excel file with proper Response import
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in export_single_table: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")

def parse_markdown_table_to_dataframe(markdown_content: str) -> pd.DataFrame:
    """Parse markdown table content to pandas DataFrame"""
    try:
        if not markdown_content:
            return pd.DataFrame()
        
        lines = markdown_content.strip().split('\n')
        table_lines = [line for line in lines if '|' in line and line.strip()]
        
        if len(table_lines) < 2:
            return pd.DataFrame()
        
        # Extract headers
        header_line = table_lines[0]
        headers = [cell.strip() for cell in header_line.split('|') if cell.strip()]
        
        # Find separator line index
        separator_idx = -1
        for i, line in enumerate(table_lines):
            if '---' in line or '--' in line:
                separator_idx = i
                break
        
        # Extract data rows
        data_start_idx = separator_idx + 1 if separator_idx >= 0 else 1
        data_rows = []
        
        for line in table_lines[data_start_idx:]:
            if '|' in line:
                row = [cell.strip() for cell in line.split('|') if cell.strip()]
                # Ensure row has same number of columns as headers
                while len(row) < len(headers):
                    row.append('')
                row = row[:len(headers)]
                data_rows.append(row)
        
        if not data_rows:
            return pd.DataFrame(columns=headers)
        
        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        return df
        
    except Exception as e:
        print(f"Error parsing markdown table: {e}")
        return pd.DataFrame()
