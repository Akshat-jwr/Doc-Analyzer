from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
import io
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import logging

from models.user import User
from models.chat_session import ChatType, ChatMessage, ChatSession
from services.multi_chat_service import multi_chat_service
from auth import get_current_active_user
from utils.pydantic_objectid import PyObjectId
from models.pdf import PDF

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/chat", tags=["Multi Chat System"])

# Global storage for download data (use Redis in production)
download_storage = {}

class StartChatRequest(BaseModel):
    chat_type: ChatType
    document_id: Optional[str] = None
    title: Optional[str] = None

class SendMessageRequest(BaseModel):
    session_id: str
    message: str

class GetSessionsRequest(BaseModel):
    chat_type: Optional[ChatType] = None
    document_id: Optional[str] = None
    limit: int = 20

class ChatResponseType:
    TEXT = "text"                    # General chat MD response
    ANALYSIS = "analysis"            # Analytical text response  
    TABLE_MODIFICATION = "table_modification"  # Modified table + download

# 📊 MARKDOWN TO EXCEL CONVERTER
class MarkdownToExcelConverter:
    """🚀 Convert markdown tables to Excel format"""
    
    @staticmethod
    def parse_markdown_table(markdown_content: str) -> List[List[str]]:
        """Parse markdown table into structured data"""
        try:
            lines = markdown_content.strip().split('\n')
            tables = []
            current_table = []
            
            for line in lines:
                line = line.strip()
                if '|' in line and line.startswith('|') and line.endswith('|'):
                    # This is a table row
                    current_table.append(line)
                elif current_table and ('---' in line or '--' in line):
                    # This is a separator line - skip it
                    continue
                elif current_table and not line:
                    # End of table - process it
                    if len(current_table) >= 2:  # At least header + data
                        tables.append(current_table)
                    current_table = []
                elif current_table and '|' not in line:
                    # End of table
                    if len(current_table) >= 2:
                        tables.append(current_table)
                    current_table = []
            
            # Process last table if exists
            if current_table and len(current_table) >= 2:
                tables.append(current_table)
            
            return tables
            
        except Exception as e:
            raise Exception(f"Error parsing markdown table: {e}")
    
    @staticmethod
    def convert_table_to_dataframe(table_lines: List[str]) -> pd.DataFrame:
        """Convert markdown table lines to pandas DataFrame"""
        try:
            # Parse header
            header_line = table_lines[0]
            headers = [col.strip() for col in header_line.split('|')[1:-1]]
            
            # Parse data rows (skip separator if present)
            data_rows = []
            for line in table_lines[1:]:
                if '---' in line or '--' in line:
                    continue  # Skip separator
                
                row_data = [col.strip() for col in line.split('|')[1:-1]]
                if len(row_data) == len(headers):
                    data_rows.append(row_data)
            
            # Create DataFrame
            df = pd.DataFrame(data_rows, columns=headers)
            return df
            
        except Exception as e:
            raise Exception(f"Error converting to DataFrame: {e}")
    
    @staticmethod
    def create_excel_from_markdown(markdown_content: str, title: str = "Modified Table") -> io.BytesIO:
        """🎯 MAIN: Convert markdown content to Excel file"""
        try:
            # Parse markdown tables
            tables = MarkdownToExcelConverter.parse_markdown_table(markdown_content)
            
            if not tables:
                raise Exception("No valid tables found in markdown content")
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Table Data"
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            current_row = 1
            
            # Process each table
            for table_idx, table_lines in enumerate(tables):
                try:
                    # Convert to DataFrame
                    df = MarkdownToExcelConverter.convert_table_to_dataframe(table_lines)
                    
                    # Add table title if multiple tables
                    if len(tables) > 1:
                        ws.cell(row=current_row, column=1, value=f"Table {table_idx + 1}")
                        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                        current_row += 2
                    
                    # Add headers
                    for col_idx, header in enumerate(df.columns, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    current_row += 1
                    
                    # Add data rows
                    for _, row in df.iterrows():
                        for col_idx, value in enumerate(row, 1):
                            ws.cell(row=current_row, column=col_idx, value=value)
                        current_row += 1
                    
                    # Add spacing between tables
                    if table_idx < len(tables) - 1:
                        current_row += 2
                        
                except Exception as e:
                    # If one table fails, add error message and continue
                    ws.cell(row=current_row, column=1, value=f"Error processing table {table_idx + 1}: {str(e)}")
                    current_row += 2
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            raise Exception(f"Error creating Excel file: {e}")

# 📥 ENHANCED TABLE DOWNLOAD SERVICE
class EnhancedTableDownloadService:
    """📥 Enhanced table download with Excel conversion"""
    
    @staticmethod
    def create_excel_download(table_markdown: str, download_id: str, query: str, filename: str = None) -> Dict[str, Any]:
        """Create Excel download from markdown table"""
        try:
            # Generate filename if not provided
            if not filename:
                filename = f"modified_table_{download_id}.xlsx"
            elif not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Convert markdown to Excel
            excel_buffer = MarkdownToExcelConverter.create_excel_from_markdown(
                table_markdown, 
                title=f"Modified Table - {download_id}"
            )
            
            return {
                "success": True,
                "content": excel_buffer.getvalue(),
                "filename": filename,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    @staticmethod
    def create_markdown_download(table_markdown: str, download_id: str, query: str) -> Dict[str, Any]:
        """Create markdown download (fallback)"""
        try:
            file_content = f"""# Modified Table Analysis
Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Query: {query}
Download ID: {download_id}

## Modified Table Data

{table_markdown}

---
Generated by Enhanced Analytical Chat System
"""
            
            return {
                "success": True,
                "content": file_content.encode('utf-8'),
                "filename": f"modified_table_{download_id}.md",
                "content_type": "text/markdown"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }

# 🔄 HELPER FUNCTIONS
def store_download_data(download_id: str, table_markdown: str, query: str):
    """Store download data for retrieval"""
    if not table_markdown:
        logger.error(f"Attempted to store empty table for download_id: {download_id}")
        return
    download_storage[download_id] = {
        'table_markdown': table_markdown,
        'query': query,
        'created_at': datetime.now(),
        'download_id': download_id
    }
    # Simple cleanup to prevent memory leaks
    if len(download_storage) > 100:
        oldest_key = min(download_storage.keys(), key=lambda k: download_storage[k]['created_at'])
        del download_storage[oldest_key]


async def get_table_from_recent_chats(download_id: str) -> Dict[str, Any]:
    """🔍 FALLBACK: Get table data from recent chat messages"""
    try:
        # Search recent chat messages for the download_id
        recent_messages = await ChatMessage.find().sort(-ChatMessage.timestamp).limit(50).to_list()
        
        for message in recent_messages:
            if (message.metadata and 
                isinstance(message.metadata, dict) and 
                message.metadata.get('download_id') == download_id and
                'new_table_markdown' in message.metadata):
                
                return {
                    'table_markdown': message.metadata['new_table_markdown'],
                    'query': message.content,
                    'download_id': download_id
                }
        
        return None
        
    except Exception as e:
        logger.error(f"Error retrieving table from chats: {e}")
        return None

# 📥 DOWNLOAD ENDPOINT
@router.get("/download/table/{download_id}")
async def download_modified_table(
    download_id: str, 
    format: str = "excel", # Default to excel
    current_user: User = Depends(get_current_active_user)
):
    """📥 Download modified table as Excel or Markdown"""
    try:
        table_data = download_storage.get(download_id)
        
        if not table_data:
            raise HTTPException(status_code=404, detail=f"Download ID '{download_id}' not found or has expired. Please try the modification again.")
        
        table_markdown = table_data.get('table_markdown')
        original_query = table_data.get('query', "Modified Table")
        
        if not table_markdown:
            raise HTTPException(status_code=500, detail="Stored table content is empty.")
        
        if format.lower() == "excel":
            download_result = EnhancedTableDownloadService.create_excel_download(
                table_markdown=table_markdown,
                download_id=download_id,
                query=original_query
            )
        else: # Fallback to markdown
            download_result = EnhancedTableDownloadService.create_markdown_download(
                table_markdown=table_markdown,
                download_id=download_id,
                query=original_query
            )
        
        if not download_result["success"]:
            raise HTTPException(status_code=500, detail=download_result["error"])
        
        headers = {
            'Content-Disposition': f'attachment; filename="{download_result["filename"]}"'
        }
        
        return Response(
            content=download_result["content"],
            media_type=download_result["content_type"],
            headers=headers
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in /download/table: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Download failed: {str(e)}")
    
    

# 🎯 EXISTING ENDPOINTS
@router.post("/start")
async def start_new_chat_session(
    request: StartChatRequest,
    current_user: User = Depends(get_current_active_user)
):
    """
    🆕 START NEW CHAT SESSION
    
    Choose chat type: general, analytical, or visualization
    Optionally attach to a document
    """
    try:
        # Verify document access if provided
        if request.document_id:
            document = await PDF.get(PyObjectId(request.document_id))
            if not document:
                raise HTTPException(status_code=404, detail="Document not found")
            
            if document.user_id != current_user.id:
                raise HTTPException(status_code=403, detail="Access denied to document")
        
        result = await multi_chat_service.start_new_chat_session(
            user_id=str(current_user.id),
            chat_type=request.chat_type,
            document_id=request.document_id,
            title=request.title
        )
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/sessions")
async def get_user_chat_sessions(
    chat_type: Optional[ChatType] = None,
    document_id: Optional[str] = None,
    limit: int = 20,
    current_user: User = Depends(get_current_active_user)
):
    """
    📋 GET USER'S CHAT SESSIONS
    
    Filter by chat type or document
    """
    try:
        # Verify document access if provided
        if document_id:
            document = await PDF.get(PyObjectId(document_id))
            if not document:
                raise HTTPException(status_code=404, detail="Document not found")
            
            if document.user_id != current_user.id:
                raise HTTPException(status_code=403, detail="Access denied to document")
        
        result = await multi_chat_service.get_user_chat_sessions(
            user_id=str(current_user.id),
            chat_type=chat_type,
            document_id=document_id,
            limit=limit
        )
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/message")
async def send_message(
    request: SendMessageRequest,
    current_user: User = Depends(get_current_active_user)
):
    try:
        result = await multi_chat_service.send_message(
            session_id=request.session_id,
            user_id=str(current_user.id),
            message=request.message
        )
        
        # ✅ THE FIX: Correctly handle the structured metadata from the service.
        if result.get("success") and "metadata" in result and isinstance(result["metadata"], dict):
            
            metadata = result["metadata"]
            
            # If it was a table modification, store the clean data for download.
            if metadata.get("response_type") == "table_modification" and metadata.get("download_ready"):
                store_download_data(
                    download_id=metadata["download_id"],
                    table_markdown=metadata["new_table_markdown"], # Use the clean markdown
                    query=request.message
                )
            
            # Pass the metadata to the frontend as a "frontend" object.
            # The frontend code already looks for this.
            result["frontend"] = metadata
        
        return result
        
    except Exception as e:
        logger.error(f"Error in /message endpoint: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
    


@router.delete("/sessions/{session_id}")
async def delete_chat_session(
    session_id: str,
    current_user: User = Depends(get_current_active_user)
):
    """
    🗑️ DELETE CHAT SESSION
    """
    try:
        result = await multi_chat_service.delete_chat_session(
            session_id=session_id,
            user_id=str(current_user.id)
        )
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/sessions/{session_id}/history")
async def get_session_history(
    session_id: str,
    limit: int = 50,
    current_user: User = Depends(get_current_active_user)
):
    """
    📜 GET SESSION CHAT HISTORY
    """
    try:
        # Verify session access
        session = await ChatSession.find_one(ChatSession.session_id == session_id)
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        if str(session.user_id) != str(current_user.id):
            raise HTTPException(status_code=403, detail="Access denied to session")
        
        # Get messages
        messages = await ChatMessage.find(
            ChatMessage.session_id == session_id
        ).sort(-ChatMessage.timestamp).limit(limit).to_list()
        
        messages.reverse()  # Chronological order
        
        return {
            "success": True,
            "session_id": session_id,
            "chat_type": session.chat_type,
            "title": session.title,
            "message_count": len(messages),
            "messages": [
                {
                    "role": msg.role,
                    "content": msg.content,
                    "timestamp": msg.timestamp.isoformat(),
                    "images_analyzed": msg.images_analyzed,
                    "metadata": msg.metadata
                }
                for msg in messages
            ]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Keep existing general chat endpoints for backward compatibility
@router.post("/general/message")
async def send_general_chat_message_legacy(
    message: dict,
    current_user: User = Depends(get_current_active_user)
):
    """
    💬 LEGACY: General chat message (backward compatibility)
    """
    try:
        # Convert to new format
        chat_type = ChatType.GENERAL
        document_id = message.get("document_id")
        user_message = message.get("message")
        conversation_id = message.get("conversation_id", "default")
        
        # Check if session exists for this conversation
        legacy_session_id = f"{current_user.id}_{conversation_id}_{document_id if document_id else 'standalone'}"
        session = await ChatSession.find_one(ChatSession.session_id == legacy_session_id)
        
        if not session:
            # Create new session for legacy support
            session_result = await multi_chat_service.start_new_chat_session(
                user_id=str(current_user.id),
                chat_type=chat_type,
                document_id=document_id,
                title=f"Legacy General Chat - {conversation_id}"
            )
            session_id = session_result["session_id"]
        else:
            session_id = session.session_id
        
        # Send message
        result = await multi_chat_service.send_message(
            session_id=session_id,
            user_id=str(current_user.id),
            message=user_message
        )
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
